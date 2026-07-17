import io
from datetime import date, datetime

import openpyxl
from pydantic import ValidationError

from app.schemas import WineCreate, WineImportError


SHEET_NAME = "WineList"

# 配布テンプレート(scripts/generate_wine_import_template.py)の列名との対応
COLUMN_TO_FIELD: dict[str, str] = {
    "No": "original_no",
    "受注日": "order_date",
    "種類": "wine_type",
    "スタイル": "style_type",
    "ワイン名": "name",
    "ワイン名カナ": "name_kana",
    "生産国": "country",
    "生産者": "producer",
    "品種": "grape_variety",
    "Vintage": "vintage",
    "サイズ": "size",
    "希望小売価格": "retail_price",
    "仕入価格(税抜/本)": "purchase_price",
    "本数": "quantity",
    "売価": "sale_price",
    "保管場所": "location",
    "コメント": "comment",
    "AI確認ステータス": "ai_check_status",
}

# 数値化できない場合に行ごとエラーにする項目(在庫数に直結するため)
STRICT_INT_FIELDS = {
    "quantity",
}

# 数値化できない場合はエラーにせずNoneとして扱う項目
# (「NV」「オープン」など業界慣習上の非数値表記が実データに現れるため)
OPTIONAL_INT_FIELDS = {
    "original_no",
    "vintage",
    "retail_price",
    "purchase_price",
    "sale_price",
}


def _parse_optional_str(value: object) -> str | None:
    if value is None:
        return None

    text = str(value).strip()

    return text or None


def _parse_optional_int(value: object) -> int | None:
    if value is None:
        return None

    if isinstance(value, str):
        text = value.strip()

        if not text:
            return None

        value = text

    try:
        return int(float(value))
    except (TypeError, ValueError) as exc:
        raise ValueError(f"数値として解釈できません: {value!r}") from exc


def _parse_optional_date(value: object) -> date | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()

    if not text:
        return None

    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError(
            f"日付として解釈できません(YYYY-MM-DD形式): {value!r}",
        ) from exc


def parse_import_workbook(
    content: bytes,
) -> tuple[list[tuple[int, WineCreate]], list[WineImportError]]:
    """
    アップロードされたxlsxの内容をパースし、行ごとにWineCreateへ変換する。

    パース・バリデーションに失敗した行はスキップしてエラー一覧に積み、
    成功した行だけを(スプレッドシート上の行番号, WineCreate)として返す。
    """

    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(content),
            data_only=True,
        )
    except Exception as exc:
        raise ValueError(
            "Excelファイルとして読み込めませんでした。"
            "xlsx形式のファイルを指定してください。",
        ) from exc

    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"シート '{SHEET_NAME}' が見つかりません。")

    sheet = workbook[SHEET_NAME]
    header = [cell.value for cell in sheet[1]]

    column_indexes: dict[str, int] = {}
    for column_name, field_name in COLUMN_TO_FIELD.items():
        if column_name in header:
            column_indexes[field_name] = header.index(column_name)

    if "name" not in column_indexes:
        raise ValueError(
            "必須列 'ワイン名' が見つかりません。テンプレートの列名を"
            "変更していないか確認してください。",
        )

    valid_rows: list[tuple[int, WineCreate]] = []
    errors: list[WineImportError] = []

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if all(
            value is None or str(value).strip() == ""
            for value in row
        ):
            continue

        data: dict[str, object] = {}
        row_error: str | None = None

        for field_name, index in column_indexes.items():
            raw_value = row[index] if index < len(row) else None

            try:
                if field_name in STRICT_INT_FIELDS:
                    data[field_name] = _parse_optional_int(raw_value)
                elif field_name in OPTIONAL_INT_FIELDS:
                    try:
                        data[field_name] = _parse_optional_int(raw_value)
                    except ValueError:
                        # NV、オープン価格など非数値の慣習表記はエラーにせず
                        # 空欄として扱う(行自体は登録を続ける)。
                        data[field_name] = None
                elif field_name == "order_date":
                    data[field_name] = _parse_optional_date(raw_value)
                else:
                    data[field_name] = _parse_optional_str(raw_value)
            except ValueError as exc:
                row_error = f"{field_name}: {exc}"
                break

        if row_error:
            errors.append(
                WineImportError(row=row_number, message=row_error),
            )
            continue

        if data.get("quantity") is None:
            data["quantity"] = 0

        try:
            wine_create = WineCreate(**data)
        except ValidationError as exc:
            first_error = exc.errors()[0]
            field = (
                first_error["loc"][0]
                if first_error["loc"]
                else "?"
            )
            errors.append(
                WineImportError(
                    row=row_number,
                    message=f"{field}: {first_error['msg']}",
                ),
            )
            continue

        valid_rows.append((row_number, wine_create))

    return valid_rows, errors

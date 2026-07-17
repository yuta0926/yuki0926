"""
旧フォーマットのワインリストExcel(受注日・ワイン名・生産者...の列を持つ台帳形式)を、
現行DBスキーマ(app/models.py の Wine)の項目順に合わせて並び替え・列名を統一する。

使い方:
    python scripts/reformat_wine_excel.py <入力xlsxパス> <出力xlsxパス>

移行内容:
- 空白のまま放置されていた未使用列(旧フォーマットの18〜27列目)を削除
- 列名をDBフィールド名に対応する形に統一
- DBには存在するがこの旧フォーマットには無かった項目(管理コード・予約本数)を
  空欄の新規列として追加(image_urlは一括インポート対象外のため含めない)
- AI確認ステータスの値を、現行アプリのAiCheckStatusBadgeが認識する語彙
  (確認済み・要確認・要修正のいずれか、それ以外は空欄)に変換する
"""

import sys
from pathlib import Path

import openpyxl


# 旧ヘッダー名 -> 新ヘッダー名 のマッピング(Noneのものは新規追加列)
COLUMN_MAPPING: list[tuple[str | None, str]] = [
    ("No", "No"),
    ("受注日", "受注日"),
    ("種類（赤・白・オレンジ・ロゼ）", "種類"),
    ("Classic/ナチュール", "スタイル"),
    ("ワイン名", "ワイン名"),
    ("ワイン名カナ", "ワイン名カナ"),
    ("国", "生産国"),
    ("生産者", "生産者"),
    ("品種", "品種"),
    ("Vintage", "Vintage"),
    ("サイズ（Bottle、HalfBottle、Glass）", "サイズ"),
    ("希望小売価格", "希望小売価格"),
    ("納価(税抜)/本", "仕入価格(税抜/本)"),
    ("本数", "本数"),
    ("売価", "売価"),
    ("場所", "保管場所"),
    (None, "管理コード"),
    (None, "予約本数"),
    ("コメント", "コメント"),
    ("AI確認フラグ", "AI確認ステータス"),
]


# 旧AI確認フラグの語彙 -> 現行アプリ(AiCheckStatusBadge)が認識する語彙
# 上記3種類以外の値は空欄(未確認扱い)にする
AI_STATUS_MAPPING: dict[str, str] = {
    "確認済み": "確認済み",
    "要確認": "要確認",
    "要修正": "要修正",
    "要補完": "要確認",
}


def reformat(input_path: Path, output_path: Path) -> int:
    source_wb = openpyxl.load_workbook(input_path, data_only=True)
    source_ws = source_wb["WineList"]

    source_header = [cell.value for cell in source_ws[1]]

    column_indexes: list[int | None] = []
    for old_name, _ in COLUMN_MAPPING:
        if old_name is None:
            column_indexes.append(None)
            continue
        column_indexes.append(source_header.index(old_name))

    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    output_ws.title = "WineList"

    output_ws.append([new_name for _, new_name in COLUMN_MAPPING])

    ai_status_column = [
        new_name for _, new_name in COLUMN_MAPPING
    ].index("AI確認ステータス")

    row_count = 0
    for row in source_ws.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in row):
            continue

        new_row = [
            row[index] if index is not None else None
            for index in column_indexes
        ]

        new_row[ai_status_column] = AI_STATUS_MAPPING.get(
            new_row[ai_status_column],
        )

        output_ws.append(new_row)
        row_count += 1

    output_wb.save(output_path)
    return row_count


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(
            "使い方: python scripts/reformat_wine_excel.py "
            "<入力xlsxパス> <出力xlsxパス>",
        )
        raise SystemExit(1)

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    count = reformat(input_path, output_path)
    print(f"{count}件のデータ行を書き出しました: {output_path}")

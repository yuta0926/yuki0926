"""
ワイン一括登録用の配布テンプレートxlsxを生成する。

現行DBスキーマ(app/models.py の Wine、app/schemas.py の WineCreate)に
対応する項目のみを含む。management_code・reserved_quantityは登録後に
管理画面側で設定する内部運用項目のため、このテンプレートには含めない。

使い方:
    python scripts/generate_wine_import_template.py <出力xlsxパス>
"""

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# (列名, 必須かどうか, 補足)
TEMPLATE_COLUMNS: list[tuple[str, bool, str]] = [
    ("No", False, "旧管理台帳の通し番号。無ければ空欄可"),
    ("受注日", False, "YYYY-MM-DD形式(例: 2026-07-17)"),
    ("種類", False, "赤 / 白 / オレンジ / ロゼ のいずれか"),
    ("スタイル", False, "Classic / ナチュール のいずれか"),
    ("ワイン名", True, "必須。空欄の行はインポート時にエラーになります"),
    ("ワイン名カナ", False, ""),
    ("生産国", False, ""),
    ("生産者", False, ""),
    ("品種", False, ""),
    ("Vintage", False, "4桁の西暦(例: 2020)。NVの場合は空欄"),
    ("サイズ", False, "Bottle / HalfBottle / Glass など"),
    ("希望小売価格", False, "半角数字(税込)"),
    ("仕入価格(税抜/本)", False, "半角数字"),
    ("本数", False, "半角数字。空欄の場合は0本として登録されます"),
    ("売価", False, "半角数字。顧客向け画面にはこの価格のみ表示されます"),
    ("保管場所", False, "例: 横浜、駒沢"),
    ("コメント", False, "社内向けメモ(顧客には表示されません)"),
    ("AI確認ステータス", False, "確認済み / 要確認 / 要修正 のいずれか。空欄可"),
]

EXAMPLE_ROW = [
    "1",
    "2026-07-17",
    "赤",
    "Classic",
    "シャトー・サンプル",
    "シャトーサンプル",
    "フランス",
    "ドメーヌ・サンプル",
    "カベルネ・ソーヴィニヨン",
    "2020",
    "Bottle",
    "6000",
    "3900",
    "12",
    "6500",
    "横浜",
    "記入例の行です。実際のデータ入力時はこの行を削除してください。",
    "要確認",
]


def build_template(output_path: Path) -> None:
    workbook = openpyxl.Workbook()

    wine_list_sheet = workbook.active
    wine_list_sheet.title = "WineList"

    header = [name for name, _, _ in TEMPLATE_COLUMNS]
    wine_list_sheet.append(header)

    header_font = Font(bold=True)
    for cell in wine_list_sheet[1]:
        cell.font = header_font

    for index, (_, is_required, _) in enumerate(TEMPLATE_COLUMNS, start=1):
        if is_required:
            wine_list_sheet.cell(row=1, column=index).fill = PatternFill(
                start_color="FFF3E0",
                end_color="FFF3E0",
                fill_type="solid",
            )

    for column_cells in wine_list_sheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        column_letter = column_cells[0].column_letter
        wine_list_sheet.column_dimensions[column_letter].width = max(
            10, min(max_length + 2, 40),
        )

    instructions_sheet = workbook.create_sheet("入力ルール")
    instructions_sheet.append(["列名", "必須", "説明"])

    for cell in instructions_sheet[1]:
        cell.font = header_font

    for name, is_required, note in TEMPLATE_COLUMNS:
        instructions_sheet.append([name, "必須" if is_required else "任意", note])

    instructions_sheet.append([])
    instructions_sheet.append([
        "記入例は「WineList」シートの2行目にあります。"
        "実際のデータを入力する際は、その行を削除するか上書きしてください。",
    ])

    for column_cells in instructions_sheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        column_letter = column_cells[0].column_letter
        instructions_sheet.column_dimensions[column_letter].width = max(
            10, min(max_length + 2, 60),
        )

    for row in instructions_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

    wine_list_sheet.append(EXAMPLE_ROW)
    for cell in wine_list_sheet[2]:
        cell.font = Font(italic=True, color="9E9E9E")

    workbook.save(output_path)


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(
            "使い方: python scripts/generate_wine_import_template.py "
            "<出力xlsxパス>",
        )
        raise SystemExit(1)

    output_path = Path(sys.argv[1])
    build_template(output_path)
    print(f"テンプレートを書き出しました: {output_path}")
